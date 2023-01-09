from tkinter import *
from tkinter import messagebox
import random
from openpyxl import load_workbook

root=Tk()
root.title("가게 키오스크 주문창")
root.geometry("800x700+500+200")


wall=PhotoImage(file="곰 세마리1.png")
wall_label=Label(root, image=wall)
wall_label.place(x=0, y=0)

menu=PhotoImage(file="메뉴판4.png")


def 테이블번호():
    table=["1", "2", "3", "4", "5", "6"]
    if random.choice(table)=="1":
        la3.configure(image=photo3)
        messagebox.showinfo("테이블 번호", "현재 앉아계신 테이블은 1번 테이블 입니다.")
    elif random.choice(table)=="2":
        la3.configure(image=photo4)
        messagebox.showinfo("테이블 번호", "현재 앉아계신 테이블은 2번 테이블 입니다.")
    elif random.choice(table)=="3":
        la3.configure(image=photo5)
        messagebox.showinfo("테이블 번호", "현재 앉아계신 테이블은 3번 테이블 입니다.")
    elif random.choice(table)=="4":
        la3.configure(image=photo6)
        messagebox.showinfo("테이블 번호", "현재 앉아계신 테이블은 4번 테이블 입니다.")
    elif random.choice(table)=="5":
        la3.configure(image=photo7)
        messagebox.showinfo("테이블 번호", "현재 앉아계신 테이블은 5번 테이블 입니다.")
    elif random.choice(table)=="6":
        la3.configure(image=photo8)
        messagebox.showinfo("테이블 번호", "현재 앉아계신 테이블은 6번 테이블 입니다.")

def 메뉴판():
    la3.configure(image=menu)

def 주문():
    messagebox.showinfo("알림", "주문 완료, 잠시만 기다려주세요")


def 큐알코드():
    la3.configure(image=photo2)
    messagebox.showinfo("큐알코드", "QR 코드를 인식하여 폼을 작성해주세요.")
    xl_file = '접종 엑셀.xlsx'

    wb = load_workbook(filename = xl_file)

    print('sheet count: ', len(wb.sheetnames))

    for sheet in wb.worksheets:    
        print('[ Sheet name: {} ]'.format(sheet.title))
        for row in sheet.rows:
            for idx, cell in enumerate(row):         
                if idx != 0:
                    print('\t', end='')
                if cell.value is None:
                    continue
                print(cell.value, end='') 
            print('')
        print('')

    wb.close()
    
photo1=PhotoImage(file="큐알코드 자리.png")
photo2=PhotoImage(file="주인 큐알코드.png")
photo3=PhotoImage(file="1번.png")
photo4=PhotoImage(file="2번.png")
photo5=PhotoImage(file="3번.png")
photo6=PhotoImage(file="4번.png")
photo7=PhotoImage(file="5번.png")
photo8=PhotoImage(file="6번.png")

la1=Label(root, text="어서오세요. 00 음식점입니다.", background="lightsalmon", font="고딕 20 bold")
la2=Label(root, text="팀명: 공(대) 세 마리", font="고딕 10 bold")
la3=Label(root, image=photo1)

bt1=Button(root, text="메뉴", bg="salmon",command=메뉴판, width=9, font="고딕 18 bold")
orderbutton=Button(root, text="주문", bg="salmon",command=주문, width=9, font="고딕 18 bold")
bt2=Button(root, text="테이블 번호 확인", bg="khaki", command=테이블번호, width=18, font="고딕 18 bold")
bt3=Button(root, text="QR 코드 촬영", bg="lightgreen", command=큐알코드, width=18, font="고딕 18 bold")

la1.place(x=230, y=50)
la2.place(x=650, y=650)
la3.place(x=70, y=150)
bt1.place(x=450, y=150)
orderbutton.place(x=600,y=150)
bt2.place(x=450, y=260)
bt3.place(x=450, y=360)


root.mainloop()
